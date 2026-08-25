#!/usr/bin/env python3
"""
KULALILAR afiş listesi şablonunu üretir.

    python3 scripts/afis-sablonu.py

Çıktı: public/sablon/kulalilar-afis-sablonu.xlsx

Sütun adları src/lib/stok-import.ts içindeki COLUMN_HINTS ile eşleşir.
Adları değiştirirsen orayı da güncelle — okuyucu başlığı adından buluyor,
sütun sırasından değil, yani sıra serbest ama ad önemli.
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# (başlık, genişlik, grup, açıklama)
#   grup: zorunlu | urun | sayfa | kampanya
COLUMNS = [
    ("SAYFA",           7,  "zorunlu",  "Sayfa numarası. Aynı numaradaki satırlar tek afişe girer."),
    ("TİP",             10, "urun",     "Boş = ürün. KAMPANYA = kampanya sayfası. HEDİYE = kampanya sayfasındaki hediye ürün."),
    ("ÜRÜN ADI",        38, "zorunlu",  "Afişte yazacak ad. Katalogla eşleşme de bu addan yapılır."),
    ("EBAT",            10, "zorunlu",  "60x120 · 7.5x30. Boş bırakırsan addan okunmaya çalışılır."),
    ("YÜZEY",           13, "urun",     "FLP · SEMİ LAPP. · MAT"),
    ("KALİTE",          9,  "urun",     "1. veya END. Çift stok kullanıyorsan bu sütun yok sayılır."),
    ("REC",             6,  "urun",     "Rektifiyeli ise E, değilse boş."),
    ("STOK",            10, "zorunlu",  "m². 1.240 veya 1240 fark etmez."),
    ("FİYAT",           9,  "zorunlu",  "TL, KDV hariç. Tam sayı."),
    ("STOK 2",          10, "urun",     "Çift stok: END. stoğu. Doluysa afişte iki satır basılır."),
    ("FİYAT 2",         10, "urun",     "Çift stok: END. fiyatı."),
    ("AFİŞ ÜRÜNÜ",      26, "urun",     "Yanlış fotoğraf geliyorsa katalogdaki doğru adı buraya yaz."),
    ("ÜRÜN SAYISI",     13, "sayfa",    "Sayfada kaç ürün olsun (1-4). Boşsa satır sayısından hesaplanır."),
    ("ZEMİN",           15, "sayfa",    "Sayfanın zemin rengi. Boşsa stüdyoda seçili olan kullanılır."),
    ("MARKA",           18, "sayfa",    "Afişin üstünde yazan marka."),
    ("SEVK YERİ",       20, "sayfa",    "PANCAR DEPO · SÖKE FABRİKA SEVK · BOZÖYÜK SEVK — listede olmayan bir yer de yazabilirsin."),
    ("KAMPANYA ÜST",    20, "kampanya", 'Küçük üst satır: "Tamamını alana"'),
    ("KAMPANYA BAŞLIK", 28, "kampanya", 'Büyük başlık: "1 palet 10x20 hediye"'),
    ("KAMPANYA METİN",  34, "kampanya", "Açıklama paragrafı. Boş bırakılabilir."),
    ("KAMPANYA NOT",    34, "kampanya", "En altta küçük punto: koşullar, son tarih."),
]

GROUP_FILL = {
    "zorunlu":  "1F2937",   # koyu — doldurulması şart
    "urun":     "4B5563",
    "sayfa":    "6B7280",
    "kampanya": "92400E",   # kampanya sütunları ayrı okunsun
}

GROUNDS = [
    "NÖTR KÂĞIT", "NÖTR AÇIK", "NÖTR ORTA", "NÖTR KOYU", "NÖTR SİYAH",
    "SICAK KÂĞIT", "SICAK AÇIK", "SICAK ORTA", "SICAK KOYU", "SICAK SİYAH",
    "ZEYTİN KÂĞIT", "ZEYTİN AÇIK", "ZEYTİN ORTA", "ZEYTİN KOYU", "ZEYTİN SİYAH",
]

# Listesi olan ama listeyle SINIRLI OLMAYAN sütunlar: öneri sunulur,
# başka bir değer yazmak engellenmez.
FREE_TEXT = {"SEVK YERİ"}

LISTS = {
    "TİP":         ["ÜRÜN", "KAMPANYA", "HEDİYE"],
    "YÜZEY":       ["FLP", "SEMİ LAPP.", "MAT"],
    "KALİTE":      ["1.", "END."],
    "REC":         ["E"],
    "ÜRÜN SAYISI": ["1", "2", "3", "4"],
    "ZEMİN":       GROUNDS,
    "SEVK YERİ":   ["PANCAR DEPO", "SÖKE FABRİKA SEVK", "BOZÖYÜK SEVK"],
}

# Örnek satırlar. SAYFA sütunu "ÖRNEK" olduğu için içe aktarmada atlanır —
# silmeyi unutsan bile afişe girmezler.
EXAMPLES = [
    ["ÖRNEK", "", "SG CIPOLLINO WHITE", "60x120", "MAT", "END.", "", "216", "425",
     "", "", "", "3", "SICAK ORTA", "GÜRAL SERAMİK", "PANCAR DEPO", "", "", "", ""],
    ["ÖRNEK", "", "SG FIROZA FULL LAP", "60x120", "FLP", "", "E", "1.231,2", "335",
     "480", "295", "", "", "", "", "", "", "", "", ""],
    ["ÖRNEK", "", "MARFIL ROSSO", "60x120", "", "1.", "", "240", "350",
     "", "", "marfil-rosso", "", "", "", "", "", "", "", ""],
    ["ÖRNEK", "KAMPANYA", "", "", "", "", "", "", "",
     "", "", "", "", "SICAK SİYAH", "GÜRAL SERAMİK", "", "Tamamını alana",
     "1 palet 10x20 hediye", "Kampanya boyunca 60x120 alımlarında geçerlidir.",
     "Bir palet 96 m² · Stoklarla sınırlıdır"],
    ["ÖRNEK", "HEDİYE", "BEYAZ PARLAK DUVAR", "10x20", "", "", "", "96", "0",
     "", "", "", "", "", "", "", "", "", "", ""],
]

THIN = Side(style="thin", color="D4D4D8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def build(path: str) -> None:
    wb = Workbook()

    # ---- 1. sayfa: AFİŞ (okunan sayfa budur, ilk sırada olmalı) ----
    ws = wb.active
    ws.title = "AFİŞ"

    for i, (title, width, group, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=title)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=GROUP_FILL[group])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30

    grey = Font(color="9CA3AF", italic=True, size=10)
    for r, row in enumerate(EXAMPLES, start=2):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = grey
            c.border = BORDER

    ws.freeze_panes = "C2"

    # Açılır listeler — yazım hatası riskini sıfırlar.
    last = len(EXAMPLES) + 400
    by_title = {t: i for i, (t, *_rest) in enumerate(COLUMNS, start=1)}
    for title, values in LISTS.items():
        col = get_column_letter(by_title[title])
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(values) + '"',
            allow_blank=True,
            showDropDown=False,
        )
        # SEVK YERİ ve MARKA gibi alanlar serbest metin: liste yalnız öneri,
        # yeni bir depo adı yazmak engellenmemeli.
        if title in FREE_TEXT:
            dv.showErrorMessage = False
        else:
            dv.error = "Bu sütunda yalnız listedeki değerler kullanılabilir."
            dv.errorTitle = "Geçersiz değer"
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last}")

    # ---- 2. sayfa: KILAVUZ ----
    gd = wb.create_sheet("KILAVUZ")
    gd.column_dimensions["A"].width = 20
    gd.column_dimensions["B"].width = 12
    gd.column_dimensions["C"].width = 86

    gd["A1"] = "KULALILAR AFİŞ LİSTESİ — KILAVUZ"
    gd["A1"].font = Font(bold=True, size=14)
    gd.merge_cells("A1:C1")

    notes = [
        "",
        "Nasıl çalışır",
        "· Aynı SAYFA numarasına yazdığın satırlar tek afiş olur. 4'ten fazlaysa kendiliğinden bölünür.",
        "· Ürün adı katalogdaki fotoğrafla otomatik eşleşir. Yanlış eşleşirse stüdyodaki önizlemede",
        "  düzeltirsin; düzeltme hatırlanır, bir dahaki dosyada tekrar sorulmaz.",
        "· İstersen AFİŞ ÜRÜNÜ sütununa doğru fotoğrafın adını yazıp baştan sabitleyebilirsin.",
        "· Fotoğrafı olmayan ürün yine sayfaya girer — yalnız görseli boş kalır ve çekim listesine düşer.",
        "· SAYFA sütununda ÖRNEK yazan satırlar okunmaz. Aşağıdaki örnekleri silmek zorunda değilsin.",
        "",
        "Sayfa ayarları (ÜRÜN SAYISI · ZEMİN · MARKA · SEVK YERİ)",
        "· Bir sayfanın herhangi bir satırına yazman yeterli, tüm sayfaya uygulanır.",
        "· Boş bırakırsan stüdyoda o an seçili olan değer kullanılır.",
        "",
        "Kampanya sayfası",
        "· TİP sütununa KAMPANYA yaz; o sayfa ürün afişi yerine kampanya sayfası olur.",
        "· Başlık ve metinleri aynı satırdaki KAMPANYA * sütunlarına yaz.",
        "· Hediye ürünleri aynı SAYFA numarasına TİP = HEDİYE olarak ekle.",
        "",
        "Çift stok",
        "· STOK 2 ve FİYAT 2 doluysa afişte iki satır basılır: 1. KALİTE ve END.",
        "· Bu durumda KALİTE sütunu yok sayılır, ikisi de yazılır.",
        "",
        "Sayı yazımı",
        "· 1.240 · 1240 · 1.240,5 hepsi doğru okunur. Fiyat tam sayıya yuvarlanır.",
        "",
    ]
    r = 2
    for line in notes:
        gd.cell(row=r, column=1, value=line)
        if line and not line.startswith(("·", " ")):
            gd.cell(row=r, column=1).font = Font(bold=True, size=11)
        gd.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    r += 1
    gd.cell(row=r, column=1, value="SÜTUNLAR").font = Font(bold=True, size=12)
    r += 1
    head = ["Sütun", "Zorunlu", "Ne yazılır"]
    for i, h in enumerate(head, start=1):
        c = gd.cell(row=r, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F2937")
        c.border = BORDER
    r += 1
    for title, _w, group, desc in COLUMNS:
        gd.cell(row=r, column=1, value=title).font = Font(bold=True, size=10)
        gd.cell(row=r, column=2, value="evet" if group == "zorunlu" else "hayır")
        gd.cell(row=r, column=3, value=desc)
        for i in (1, 2, 3):
            gd.cell(row=r, column=i).border = BORDER
            gd.cell(row=r, column=i).alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    r += 1
    gd.cell(row=r, column=1, value="ZEMİN DEĞERLERİ").font = Font(bold=True, size=12)
    r += 1
    for g in GROUNDS:
        gd.cell(row=r, column=1, value=g)
        r += 1

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"yazıldı: {path}")


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "public/sablon/kulalilar-afis-sablonu.xlsx"
    build(out)
